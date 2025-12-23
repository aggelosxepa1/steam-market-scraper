from __future__ import annotations
from typing import List, Tuple, Optional, Any
import re
import asyncio
import aiohttp
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.common.exceptions import WebDriverException, TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import urllib3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

item_type: str | None = None
min_price: int | None = None
max_price: int | None = None

while True:
    try:
        if item_type is None:
            temp_input = input("Enter the item type e.g. gloves, knife: ")
            if len(temp_input) < 2 or temp_input.isdigit():
                raise ValueError("Invalid item type")
            item_type = temp_input

        if min_price is None:
            temp_min = int(input("Minimum price (USD): "))
            if temp_min < 1:
                raise ValueError("Invalid minimum price")
            min_price = temp_min

        if max_price is None:
            temp_max = int(input("Maximum price (USD): "))
            if temp_max <= min_price:
                raise ValueError("Maximum price must be greater than minimum price")
            max_price = temp_max
        break
    except (TypeError, ValueError) as e:
        print(e)
        
cleaned_user_input: str = re.sub(r"[^\w]", " ", item_type.lower()).strip()

file_path: str = "steam_market.xlsx"
not_include: List[str] = ["Key", "Case", "Capsule", "Package", "Music Kit", "Sticker", "Patch"]
proxy_index: int = 0
timeout: int = 20

def load_proxies(file_path: str = "proxies.txt") -> List[str]:
    try:
        with open(file_path, "r") as file:
            return [line.strip() for line in file if line.strip()]
    except FileNotFoundError as e:
        print("Proxy file not found.")
        print(e)
        return []

raw_proxies: List[str] = load_proxies()
if not raw_proxies:
    print("No proxies found. Exiting.")
    exit()

async def validate_proxy_async(session: aiohttp.ClientSession, proxy: str) -> bool:
    try:
        proxy_url = f"http://{proxy}"
        async with session.get(
            "https://store.steampowered.com/",
            proxy=proxy_url,
            timeout=5
        ) as resp:
            return resp.status == 200
    except Exception:
        return False

def get_driver_with_proxy(proxy: str) -> WebDriver:
    options = webdriver.ChromeOptions()
    options.add_argument(f"--proxy-server=http://{proxy}")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("detach", True)
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.minimize_window()
    return driver

async def get_working_driver(
    proxy_list: List[str], start_index: int
) -> Tuple[WebDriver, int]:
    async with aiohttp.ClientSession() as session:
        while True:
            for idx in range(start_index, len(proxy_list)):
                current_proxy = proxy_list[idx]
                print(f"Testing proxy {idx+1}/{len(proxy_list)}: {current_proxy}")
                if await validate_proxy_async(session, current_proxy):
                    try:
                        driver = get_driver_with_proxy(current_proxy)
                        print(f"Using proxy: {current_proxy}")
                        driver.close()
                        return driver, idx
                    except Exception as e:
                        print(f"Selenium failed with proxy {current_proxy}: {e}")
            start_index = 0

def rotate_proxy(driver: Optional[WebDriver],
proxy_index: int,
proxy_list: List[str]
) -> Tuple[WebDriver, int]:
    try:
        if driver:
            driver.quit()
    except Exception:
        pass
    proxy_index += 1
    if proxy_index >= len(proxy_list):
        proxy_index = 0
    return asyncio.run(get_working_driver(proxy_list, proxy_index))

def clean_empty_rows(sheet: Any) -> None:
    for row in range(sheet.max_row, 0, -1):
        if all((cell.value is None or str(cell.value).strip() == "") for cell in sheet[row]):
            sheet.delete_rows(row)

def append_to_excel(file_path: str, row_data: List[Any]) -> None:
    try:
        if not os.path.exists(file_path):
            df_new = pd.DataFrame([row_data], columns=["Item Name", "Skin Quality", "Buy Order Price", "Item URL"])
            df_new.to_excel(file_path, index=False)
            wb = load_workbook(file_path)
            ws = wb.active
            url_cell = ws.cell(row=2, column=4)
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0000FF", underline="single")
            wb.save(file_path)
        else:
            book = load_workbook(file_path)
            sheet = book.active
            clean_empty_rows(sheet)
            start_row = sheet.max_row + 1
            for r in range(2, sheet.max_row + 2):
                if sheet.cell(r, 1).value in (None, ""):
                    start_row = r
                    break

            for col_index, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=start_row, column=col_index, value=value)
                if col_index == 4:
                    cell.hyperlink = value
                    cell.font = Font(color="0000FF", underline="single")

            book.save(file_path)

    except PermissionError:
        print(f"Permission denied: Is the Excel file open? -> {file_path}")
    except Exception as e:
        print(f"Error appending to Excel: {e}")

driver, proxy_index = asyncio.run(get_working_driver(raw_proxies, proxy_index))

def fetch_single_item_selenium(url: str) -> Optional[List[Any]]:
    global proxy_index
    driver = get_driver_with_proxy(raw_proxies[proxy_index])
    try:
        driver.get(url)
        soup_check = BeautifulSoup(driver.page_source, "html.parser")
        WebDriverWait(driver, timeout).until(lambda d: d.execute_script('return document.readyState') == 'complete')
        too_many_requests = soup_check.find("p", class_="sectionText",
                                  string=lambda t: t and "An error was encountered while processing your request:" in t)
        if too_many_requests:
            print("Too many requests on page 1, rotating proxy...")
            driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CLASS_NAME, "hover_item_name"))
        )
        title = soup.find("h1", class_="hover_item_name").text.split()
        if not any(word in not_include for word in title):
            span_tags = soup.find_all("span", class_="market_commodity_orders_header_promote")
            price_str = span_tags[1].text.split("$")[1].split(".")[0]
            price = int(price_str)
            if min_price <= price <= max_price:
                quality = soup.find("div", class_="descriptor").text
                item_name = " ".join(title)
                return [item_name, quality, price, url]
    except (WebDriverException, IndexError, AttributeError, ValueError, urllib3.exceptions.ReadTimeoutError) as e:
        print(f"Error processing item: {e}")
        driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
    finally:
        driver.quit()
    return None

async def get_item_info(a_tags: List[Any]) -> None:
    tasks = []
    for tag in a_tags:
        url = tag["href"]
        tasks.append(asyncio.to_thread(fetch_single_item_selenium, url))
    results = await asyncio.gather(*tasks)
    results = [r for r in results if r is not None]
    for row in results:
        append_to_excel(file_path, row)

def fetch_data(proxy_index: int) -> None:
    url = f"https://steamcommunity.com/market/search?appid=730&q={cleaned_user_input}#p1_price_asc"
    driver = get_driver_with_proxy(raw_proxies[proxy_index])

    try:
        driver.get(url)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        span_pagelink = soup.find_all("span", class_="market_paging_pagelink")
        if span_pagelink:
            pages = int(span_pagelink[-1].text)
    except Exception as e:
        print(f"Error getting page count: {e}")
        driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)

    WebDriverWait(driver, timeout).until(lambda d: d.execute_script('return document.readyState') == 'complete')

    too_many_requests = soup.find("p", class_="sectionText",
                                  string=lambda t: t and "An error was encountered while processing your request:" in t)
    if too_many_requests:
        print("Too many requests on page 1, rotating proxy...")
        driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
    try:
        WebDriverWait(driver, 6).until(EC.presence_of_element_located((By.CLASS_NAME, "market_listing_row_link")))
    except TimeoutException as e:
        print(e)
        driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    a_tags = soup.find_all("a", class_="market_listing_row_link")
    driver.quit()
    asyncio.run(get_item_info(a_tags))

    for page_num in range(2, pages + 1):
        page_url = f"https://steamcommunity.com/market/search?appid=730&q={cleaned_user_input}#p{page_num}_price_asc"
        driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
        try:
            driver.get(page_url)
            WebDriverWait(driver, timeout).until(lambda d: d.execute_script('return document.readyState') == 'complete')
            soup = BeautifulSoup(driver.page_source, "html.parser")
            too_many_requests = soup.find("p", class_="sectionText",
                                        string=lambda t: t and "An error was encountered while processing your request:" in t)
            if too_many_requests:
                print(f"Too many requests on page {page_num}, rotating proxy...")
                driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
                driver.get(page_url)

            WebDriverWait(driver, timeout).until(lambda d: d.execute_script('return document.readyState') == 'complete')
            soup = BeautifulSoup(driver.page_source, "html.parser")
            WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.CLASS_NAME, "market_listing_row_link")))

        except (WebDriverException, TimeoutException, urllib3.exceptions.ReadTimeoutError) as e:
            print(e)
            driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)

        a_tags = soup.find_all("a", class_="market_listing_row_link")
        if len(a_tags) == 0:
            driver, proxy_index = rotate_proxy(driver, proxy_index, raw_proxies)
        driver.quit()
        asyncio.run(get_item_info(a_tags))

fetch_data(proxy_index)
print("Finished scraping.")

